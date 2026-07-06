"""Generate src/lib/data/module-catalog.ts from the Start Academy Excel catalogue.

This script is the single source of truth for transforming the curated
Excel sheets (Conseiller / Manager / Assistantes + diagnostic helpers)
into typed TypeScript data consumed by the Next.js app.

Usage (from the project root):

    npm run generate:catalog

Or directly:

    python3 scripts/generate-module-catalog/generate_catalog.py

Requirements: Python 3.9+ and `openpyxl` (pip3 install openpyxl).

Paths are resolved relative to this file, so the script works regardless
of the current working directory.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl  # type: ignore[import-untyped]
except ImportError as exc:
    print(
        "ERROR: openpyxl is required. Install it with: pip3 install openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


# Resolve paths from this file's location.
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent          # start-academy-diagnostic/
REPO_ROOT = PROJECT_ROOT.parent                   # Start Proposition/
EXCEL_PATH = REPO_ROOT / "data" / "catalogue-formations-start-academy.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "src" / "lib" / "data" / "module-catalog.ts"
SEED_OUTPUT_PATH = PROJECT_ROOT / "supabase" / "seed.sql"


FOUNDATION_KEYWORDS = [
    "chat gpt",
    "chatgpt",
    "gpt",
    "claude",
    "gamma",
    "notebook lm",
    "notebooklm",
    "notebook",
    "gemini",
    "workspace",
    "prompt",
    "crm",
    "canva",
]

SHEET_TO_PROFILE = {
    "Conseiller": "conseiller",
    "Manager": "manager",
    "Assistantes": "assistant",
}


def slugify(value: str) -> str:
    s = value.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def parse_duration(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lower()
    if not text:
        return None
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*h", text)
    if match:
        return float(match.group(1).replace(",", "."))
    match = re.search(r"(\d+(?:[.,]\d+)?)", text)
    if match:
        return float(match.group(1).replace(",", "."))
    return None


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def is_section_header(row) -> bool:
    """Inside each sheet, section sub-headers reuse the column-2 label 'Durée'."""
    return (
        len(row) >= 2
        and isinstance(row[1], str)
        and row[1].strip().lower() == "durée"
    )


def is_foundation(name: str, family) -> bool:
    name_lower = name.lower()
    if any(kw in name_lower for kw in FOUNDATION_KEYWORDS):
        return True
    if family and "base" in family.lower():
        return True
    return False


def normalize_for_match(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9 ]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def squish(value: str) -> str:
    return re.sub(r"\s+", "", normalize_for_match(value))


def extract_modules(workbook, sheet_name: str, profile: str):
    sheet = workbook[sheet_name]
    modules = []
    current_family = None

    for row in sheet.iter_rows(values_only=True):
        if not row or not any(c not in (None, "") for c in row):
            continue
        first = clean(row[0])
        if first is None:
            continue
        if is_section_header(row):
            current_family = first
            continue
        if current_family is None:
            continue

        level_raw = clean(row[3]) if len(row) > 3 else None
        if isinstance(level_raw, (int, float)):
            level_value = float(level_raw)
        elif isinstance(level_raw, str):
            try:
                level_value = float(level_raw)
            except ValueError:
                level_value = None
        else:
            level_value = None

        paying_raw = clean(row[5]) if len(row) > 5 else None
        if isinstance(paying_raw, str):
            paying_lower = paying_raw.strip().lower()
            if paying_lower in ("oui", "yes", "y"):
                paying_value = True
            elif paying_lower in ("non", "no", "n"):
                paying_value = False
            else:
                paying_value = None
        else:
            paying_value = None

        module_id = f"{profile}-{slugify(current_family)}-{slugify(first)}"
        modules.append(
            {
                "id": module_id,
                "name": first,
                "family": current_family,
                "sourceSheet": sheet_name,
                "targetProfile": profile,
                "durationHours": parse_duration(clean(row[1]) if len(row) > 1 else None),
                "level": level_value,
                "tools": clean(row[4]) if len(row) > 4 else None,
                "paying": paying_value,
                "platform": clean(row[6]) if len(row) > 6 else None,
                "needIdentification": clean(row[2]) if len(row) > 2 else None,
                "isFoundationModule": is_foundation(first, current_family),
                "diagnosticSignals": [],
            }
        )
    return modules


def collect_signal_triggers(workbook):
    """Return list of (signal_phrase, [trigger_tokens]) extracted from the
    Guide diagnostic IA, Diagnostic performance and Diagnostic outils de
    base sheets."""
    triggers = []

    sheet = workbook["Guide diagnostic IA"]
    seen_header = False
    for row in sheet.iter_rows(values_only=True):
        a = clean(row[0])
        b = clean(row[1]) if len(row) > 1 else None
        if a and b and "signal entendu" in a.lower():
            seen_header = True
            continue
        if seen_header and a and b:
            if "règle" in a.lower():
                break
            tokens = [t.strip() for t in re.split(r",|/|;", b) if t.strip()]
            triggers.append((a, tokens))

    sheet = workbook["Diagnostic performance"]
    for row in sheet.iter_rows(values_only=True):
        if not row or all(c in (None, "") for c in row):
            continue
        family = clean(row[0])
        threshold = clean(row[3]) if len(row) > 3 else None
        modules_col = clean(row[4]) if len(row) > 4 else None
        if family and modules_col and family.lower() not in (
            "famille",
            "objectif",
            "outils de base",
            "règle de scoring recommandée",
        ):
            tokens = [t.strip() for t in re.split(r",|/|;", modules_col) if t.strip()]
            phrase = f"{family} — {threshold}" if threshold else family
            triggers.append((phrase, tokens))

    sheet = workbook["Diagnostic outils de base"]
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        tool = clean(row[0])
        weak_signals = clean(row[4]) if len(row) > 4 else None
        modules_col = clean(row[5]) if len(row) > 5 else None
        if tool and modules_col:
            tokens = [t.strip() for t in re.split(r",|/|;", modules_col) if t.strip()]
            phrase = f"{tool} — {weak_signals}" if weak_signals else tool
            triggers.append((phrase, tokens))

    return triggers


def attach_signals(modules, triggers):
    norm = {m["id"]: (normalize_for_match(m["name"]), squish(m["name"])) for m in modules}
    for phrase, tokens in triggers:
        for token in tokens:
            nt = normalize_for_match(token)
            st = squish(token)
            if len(nt) < 3:
                continue
            for module in modules:
                module_norm, module_squish = norm[module["id"]]
                if (
                    nt in module_norm
                    or module_norm in nt
                    or st in module_squish
                    or module_squish in st
                ):
                    if phrase not in module["diagnosticSignals"]:
                        module["diagnosticSignals"].append(phrase)


def extract_performance_families(workbook):
    sheet = workbook["Diagnostic performance"]
    out = []
    in_block = False
    for row in sheet.iter_rows(values_only=True):
        a = clean(row[0])
        b = clean(row[1]) if len(row) > 1 else None
        if a == "Famille" and b == "Questions à poser en RDV":
            in_block = True
            continue
        if in_block:
            if a is None:
                in_block = False
                continue
            if a.lower().startswith("règle"):
                in_block = False
                continue
            if a.lower() == "outils de base":
                break
            out.append(
                {
                    "family": a,
                    "questions": b,
                    "ratio": clean(row[2]) if len(row) > 2 else None,
                    "threshold": clean(row[3]) if len(row) > 3 else None,
                    "triggeredModules": clean(row[4]) if len(row) > 4 else None,
                }
            )
    return out


def extract_tool_diagnostic(workbook):
    sheet = workbook["Diagnostic outils de base"]
    out = []
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        if not row or all(c in (None, "") for c in row):
            continue
        out.append(
            {
                "tool": clean(row[0]),
                "questions": clean(row[1]) if len(row) > 1 else None,
                "level": clean(row[2]) if len(row) > 2 else None,
                "performanceImpact": clean(row[3]) if len(row) > 3 else None,
                "weakSignals": clean(row[4]) if len(row) > 4 else None,
                "moduleToTrigger": clean(row[5]) if len(row) > 5 else None,
            }
        )
    return out


def extract_signal_map(workbook):
    sheet = workbook["Guide diagnostic IA"]
    out = []
    seen_header = False
    for row in sheet.iter_rows(values_only=True):
        a = clean(row[0])
        b = clean(row[1]) if len(row) > 1 else None
        if a and b and "signal entendu" in a.lower():
            seen_header = True
            continue
        if seen_header and a and b:
            if "règle" in a.lower():
                break
            out.append({"signal": a, "modules": b})
    return out


def ts_value(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return f"{value}"
    if isinstance(value, list):
        if not value:
            return "[]"
        inner = ",\n      ".join(ts_value(item) for item in value)
        return "[\n      " + inner + ",\n    ]"
    return json.dumps(value, ensure_ascii=False)


def emit_record(record, fields) -> str:
    lines = ["  {"]
    for key in fields:
        lines.append(f"    {key}: {ts_value(record[key])},")
    lines.append("  }")
    return "\n".join(lines)


def render_typescript(data) -> str:
    module_fields = [
        "id",
        "name",
        "family",
        "sourceSheet",
        "targetProfile",
        "durationHours",
        "level",
        "tools",
        "paying",
        "platform",
        "needIdentification",
        "isFoundationModule",
        "diagnosticSignals",
    ]
    perf_fields = ["family", "questions", "ratio", "threshold", "triggeredModules"]
    tool_fields = [
        "tool",
        "questions",
        "level",
        "performanceImpact",
        "weakSignals",
        "moduleToTrigger",
    ]

    modules_ts = ",\n".join(emit_record(m, module_fields) for m in data["modules"])
    perf_ts = ",\n".join(
        emit_record(p, perf_fields) for p in data["performanceFamilies"]
    )
    tool_ts = ",\n".join(emit_record(t, tool_fields) for t in data["toolDiagnostic"])
    signals_ts = ",\n".join(
        f"  {{ signal: {json.dumps(s['signal'], ensure_ascii=False)}, "
        f"modules: {json.dumps(s['modules'], ensure_ascii=False)} }}"
        for s in data["signalMap"]
    )

    return (
        "// Auto-généré depuis data/catalogue-formations-start-academy.xlsx\n"
        "// Ne pas éditer à la main : régénérer via\n"
        "//   npm run generate:catalog\n"
        "// (cf. scripts/generate-module-catalog/generate_catalog.py).\n"
        "\n"
        "import type {\n"
        "  TrainingModule,\n"
        "  PerformanceDiagnosticFamily,\n"
        "  ToolBaseDiagnostic,\n"
        "  DiagnosticSignalMapping,\n"
        "} from \"@/types\";\n"
        "\n"
        "export const moduleCatalog: TrainingModule[] = [\n"
        f"{modules_ts},\n"
        "];\n"
        "\n"
        "export const performanceDiagnosticFamilies: PerformanceDiagnosticFamily[] = [\n"
        f"{perf_ts},\n"
        "];\n"
        "\n"
        "export const toolBaseDiagnostic: ToolBaseDiagnostic[] = [\n"
        f"{tool_ts},\n"
        "];\n"
        "\n"
        "export const diagnosticSignalMap: DiagnosticSignalMapping[] = [\n"
        f"{signals_ts},\n"
        "];\n"
    )


def sql_string(value) -> str:
    """Encode a string for a SQL literal (single-quoted, with E'' escaping)."""
    if value is None:
        return "null"
    escaped = value.replace("\\", "\\\\").replace("'", "''")
    return f"E'{escaped}'" if "\n" in value or "\\" in value else f"'{escaped}'"


def sql_value(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        if not value:
            return "ARRAY[]::text[]"
        items = ", ".join(sql_string(item) for item in value)
        return f"ARRAY[{items}]::text[]"
    return sql_string(value)


def render_seed_sql(modules) -> str:
    """Render an idempotent seed for training_modules.

    Uses ON CONFLICT to allow `supabase db reset` and re-runs."""
    columns = [
        ("id", "id"),
        ("name", "name"),
        ("family", "family"),
        ("source_sheet", "sourceSheet"),
        ("target_profile", "targetProfile"),
        ("duration_hours", "durationHours"),
        ("level", "level"),
        ("tools", "tools"),
        ("paying", "paying"),
        ("platform", "platform"),
        ("need_identification", "needIdentification"),
        ("is_foundation_module", "isFoundationModule"),
        ("diagnostic_signals", "diagnosticSignals"),
    ]
    col_names = ", ".join(c for c, _ in columns)
    update_clause = ",\n  ".join(
        f"{c} = excluded.{c}" for c, _ in columns if c != "id"
    )

    rows = []
    for module in modules:
        values = ",\n    ".join(sql_value(module[key]) for _, key in columns)
        rows.append(f"  (\n    {values}\n  )")
    values_block = ",\n".join(rows)

    return (
        "-- Auto-généré depuis data/catalogue-formations-start-academy.xlsx\n"
        "-- Ne pas éditer à la main : régénérer via `npm run generate:catalog`.\n"
        "--\n"
        "-- Ce seed est idempotent : `supabase db reset` ou un nouvel appel\n"
        "-- écrasera proprement les modules existants.\n"
        "\n"
        f"insert into public.training_modules ({col_names}) values\n"
        f"{values_block}\n"
        "on conflict (id) do update set\n"
        f"  {update_clause};\n"
    )


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        raise SystemExit(1)

    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    modules = []
    for sheet_name, profile in SHEET_TO_PROFILE.items():
        if sheet_name not in workbook.sheetnames:
            print(
                f"WARN: sheet '{sheet_name}' is missing from the Excel file",
                file=sys.stderr,
            )
            continue
        modules += extract_modules(workbook, sheet_name, profile)

    triggers = collect_signal_triggers(workbook)
    attach_signals(modules, triggers)

    data = {
        "modules": modules,
        "performanceFamilies": extract_performance_families(workbook),
        "toolDiagnostic": extract_tool_diagnostic(workbook),
        "signalMap": extract_signal_map(workbook),
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(render_typescript(data), encoding="utf-8")

    SEED_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SEED_OUTPUT_PATH.write_text(render_seed_sql(modules), encoding="utf-8")

    print(f"Wrote {OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Wrote {SEED_OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"  modules:            {len(modules)}")
    print(f"  performanceFamilies: {len(data['performanceFamilies'])}")
    print(f"  toolDiagnostic:     {len(data['toolDiagnostic'])}")
    print(f"  signalMap:          {len(data['signalMap'])}")
    print(
        f"  foundation modules: "
        f"{sum(1 for m in modules if m['isFoundationModule'])}"
    )


if __name__ == "__main__":
    main()
